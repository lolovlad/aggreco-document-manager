from .BuilderFile import BuilderFile

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill

from ..FileParser.JsonFile import JsonFile

from ..Models.Fileschame import *

from copy import deepcopy
from pathlib import Path
from re import findall


class BuilderXlsxFile(BuilderFile):
    def __init__(self, path_file: Path, file_schemas: FileSchemas):
        self.__wb: Workbook = None
        self.__active_sheet: Worksheet = None
        self.__file_schemas: FileSchemas = file_schemas
        self.__json_data_schema: dict = {}
        self.__border_style: Border = Border(left=Side(style='thin'),
                                             right=Side(style='thin'),
                                             top=Side(style='thin'),
                                             bottom=Side(style='thin'))
        self.__fill_yellow: PatternFill = PatternFill(start_color="5cb800",
                                                      fill_type="solid")

        self.__path_file: Path = path_file

    @property
    def json_data_schema(self):
        return deepcopy(self.__json_data_schema)

    def build(self):
        self.__create_workbook()

        for target_protocol in self.__file_schemas.protocols:
            self.__create_sheet(target_protocol)

            bias = 0
            last_y_cell = 0
            for table in target_protocol.tables:
                for cell in table.cells:

                    x = cell.x
                    y = cell.y + bias + last_y_cell

                    coord = (x, y)

                    cell_object = self.__get_cell(x, y)

                    if cell.is_merge:
                        self.__create_merge_cell(cell_object, coord, cell.size, cell.text)
                    elif cell.is_data:
                        self.__create_data_cell(cell_object, coord, cell.text)
                    else:
                        self.__create_cell(cell_object, cell.text)

                last_y_cell = y - 2
                bias += 10
        self.__save_file()

    def __create_workbook(self):
        self.__wb = Workbook()

    def __create_sheet(self, protocol: Protocol):
        self.__active_sheet = self.__wb.create_sheet(protocol.name)
        self.__json_data_schema[protocol.name] = {}

    def __fill_cell(self, cell):
        cell.fill = self.__fill_yellow

    def __get_cell(self, x: int, y: int):
        return self.__active_sheet.cell(row=y, column=x)

    def __get_json_data_schema(self, sheet, key):
        return self.__json_data_schema[sheet][key]

    def __is_cell_ref(self, text: str):
        coord = self.__json_data_schema[self.__active_sheet.title].get(text)
        return coord is not None

    def __create_data_cell(self, cell_object, coord: tuple, text: str):
        text = self.__get_text_in_cell(text)
        if self.__is_cell_ref(text):
            coord_ref_cell = self.__get_json_data_schema(self.__active_sheet.title, text)
            text = self.__create_ref_to_cell(coord_ref_cell)
        else:
            self.__add_to_schema(text, coord)
            self.__fill_cell(cell_object)
            text = ""
        self.__create_cell(cell_object, text)

    def __create_cell(self, cell, text: str):
        cell.value = text
        cell.border = self.__border_style

    def __create_merge_cell(self, cell, coord: tuple, size: Size, text: str):
        try:
            self.__create_cell(cell, text)

            end_x = coord[0] + size.width - 1
            end_y = coord[1] + size.height - 1

            self.__active_sheet.merge_cells(start_row=coord[1],
                                            start_column=coord[0],
                                            end_row=end_y,
                                            end_column=end_x)
        except AttributeError:
            pass

    def __add_to_schema(self, text: str, coord: tuple):
        title_sheet = self.__active_sheet.title
        self.__json_data_schema[title_sheet][self.__get_text_in_cell(text)] = coord

    def __create_ref_to_cell(self, coord: tuple) -> str:
        cell_ref = self.__get_cell(coord[0], coord[1])
        column_letter = cell_ref.column_letter
        row = cell_ref.row
        return f"={column_letter}{row}"

    def __get_text_in_cell(self, text) -> str:
        text_list = findall(r"\w+", text)
        return text_list[0]

    def __save_file(self):
        self.__wb.save(self.__path_file)
        json_file = JsonFile(self.__path_file, self.__json_data_schema)
        json_file.save_file()