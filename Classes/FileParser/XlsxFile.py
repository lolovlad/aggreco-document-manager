from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill

from .File import File
from pathlib import Path

from ..Models.Fileschame import *

from .JsonFile import JsonFile

from re import findall


class XlsxFile(File):
    def __init__(self, path_file: Path, map_data: FileSchemas):
        self.__path_file: Path = path_file
        self.__file: Workbook = Workbook()
        self.__map_data: FileSchemas = map_data
        self.__active_sheet: Worksheet = None

        self.__border_style: Border = Border(left=Side(style='thin'),
                                             right=Side(style='thin'),
                                             top=Side(style='thin'),
                                             bottom=Side(style='thin'))
        self.__fill_yellow: PatternFill = PatternFill(start_color="5cb800",
                                                      fill_type="solid")

    @property
    def map_data(self):
        return self.__map_data

    def read_file(self):
        self.__file = load_workbook(filename=self.__path_file)

    def render(self, schemas: dict):
        pass

    def is_table_schema(self, table, re: str):
        pass

    def get_all_parser_table_in_file(self) -> list[TableSchemas]:
        list_tables = []
        for protocols in self.__map_data.protocols:
            list_tables += protocols.tables
        return list_tables

    def get_list_cells(self, table):
        pass

    def get_sheet_by_name(self, name: str):
        return self.__file.get_sheet_by_name(name)

    def target_sheet_by_name(self, name: str):
        self.__active_sheet = self.__file.get_sheet_by_name(name)

    def save(self):
        self.__file.save(self.__path_file)

    def create_sheet(self, protocol: Protocol):
        self.__active_sheet = self.__file.create_sheet(protocol.name)

    def get_cell(self, x: int, y: int):
        return self.__active_sheet.cell(row=y, column=x)

    def create_cell(self, x: int, y: int, text: str):
        object_cell = self.get_cell(x, y)
        object_cell.value = text
        object_cell.border = self.__border_style

    def get_coord_cell(self, x: int, y: int) -> str:
        cell = self.get_cell(x, y)
        column_letter = cell.column_letter
        row = cell.row
        return f"{column_letter}{row}"

    def add_fill_cell(self, x: int, y: int):
        cell = self.get_cell(x, y)
        cell.fill = self.__fill_yellow

    def merge_cells(self, start_x: int, start_y: int, end_x: int, end_y: int):
        self.__active_sheet.merge_cells(start_row=start_y,
                                        start_column=start_x,
                                        end_row=end_y,
                                        end_column=end_x)

    def get_title_sheet(self) -> str:
        return self.__active_sheet.title

    def get_text_in_cell(self, text) -> str:
        text_list = findall(r"\w+", text)
        return text_list[0]
